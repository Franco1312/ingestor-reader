"""BCRA Infomondia parser plugin."""

import io
from datetime import datetime
from typing import Dict, List, Optional, Union

from openpyxl import load_workbook

from src.domain.interfaces import Parser
from src.infrastructure.utils.date_utils import to_naive
from src.infrastructure.utils.excel_utils import excel_column_to_index

CellValue = Union[datetime, str, float, int, None]
SeriesDataPoint = Dict[str, Union[str, CellValue]]


class BcraInfomondiaParser(Parser):
    """Parser for BCRA Infomondia Excel files."""

    def parse(
        self, 
        raw_data: bytes, 
        config: Dict[str, Union[str, int, Dict, List]],
        series_last_dates: Optional[Dict[str, datetime]] = None
    ) -> List[SeriesDataPoint]:
        """Parse raw Excel data according to BCRA Infomondia configuration.
        
        Args:
            raw_data: Raw bytes of the Excel file.
            config: Configuration dictionary containing parse_config.
            series_last_dates: Dictionary mapping series_code to last processed date.
            
        Returns:
            List of dictionaries containing parsed series data.
        """
        parse_config = config.get("parse_config", {})
        series_map = parse_config.get("series_map", [])
        
        workbook = self._load_workbook(raw_data)
        parsed_data: List[SeriesDataPoint] = []
        
        try:
            for series_config in series_map:
                series_code = str(series_config.get("internal_series_code", ""))
                last_date = series_last_dates.get(series_code) if series_last_dates else None
                series_data = self._extract_series(workbook, series_config, last_date)
                parsed_data.extend(series_data)
        finally:
            workbook.close()
        
        return parsed_data
    
    def _load_workbook(self, raw_data: bytes):
        """Load Excel workbook from bytes.
        
        Args:
            raw_data: Raw bytes of the Excel file.
            
        Returns:
            Loaded workbook object.
        """
        excel_file = io.BytesIO(raw_data)
        return load_workbook(excel_file, data_only=True, read_only=True)
    
    def _extract_series(
        self, 
        workbook, 
        series_config: Dict[str, Union[str, int, bool]],
        last_date: Optional[datetime] = None
    ) -> List[SeriesDataPoint]:
        """Extract data for a single series from the workbook.
        
        Args:
            workbook: OpenPyXL workbook object.
            series_config: Configuration for the series to extract.
            last_date: Last processed date for this series (filters dates <= last_date).
            
        Returns:
            List of dictionaries containing series data points.
        """
        sheet_name = series_config["sheet"]
        
        if sheet_name not in workbook.sheetnames:
            return []
        
        sheet = workbook[sheet_name]
        header_row = series_config["header_row"]
        date_col = series_config["date_col"]
        value_col = series_config["value_col"]
        drop_na = series_config.get("drop_na", False)
        skip_rows_after_header = series_config.get("skip_rows_after_header", 1)
        
        date_col_idx = excel_column_to_index(date_col)
        value_col_idx = excel_column_to_index(value_col)
        
        dates: List[Optional[CellValue]] = []
        values: List[Optional[CellValue]] = []
        
        for row in sheet.iter_rows(
            min_row=header_row + skip_rows_after_header,
            values_only=True
        ):
            date = row[date_col_idx] if date_col_idx < len(row) else None
            value = row[value_col_idx] if value_col_idx < len(row) else None
            
            if last_date and isinstance(date, datetime):
                if date <= to_naive(last_date):
                    continue
            
            dates.append(date)
            values.append(value)
        
        return self._build_series_data(
            series_config, 
            dates, 
            values, 
            drop_na
        )
    
    def _build_series_data(
        self,
        series_config: Dict[str, Union[str, int, bool]],
        dates: List[Optional[CellValue]],
        values: List[Optional[CellValue]],
        drop_na: bool
    ) -> List[SeriesDataPoint]:
        """Build series data points from dates and values.
        
        Args:
            series_config: Configuration for the series.
            dates: List of date values.
            values: List of value values.
            drop_na: Whether to drop None/NA values.
            
        Returns:
            List of dictionaries containing series data points.
        """
        series_data: List[SeriesDataPoint] = []
        
        for date, value in zip(dates, values):
            if drop_na and (date is None or value is None):
                continue
            
            series_data.append({
                "internal_series_code": str(series_config["internal_series_code"]),
                "unit": series_config.get("unit"),
                "frequency": series_config.get("frequency"),
                "obs_time": date,
                "value": value,
            })
        
        return series_data
