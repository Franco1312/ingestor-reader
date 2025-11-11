"""Pytest configuration and shared fixtures."""

import io
from datetime import datetime

import pytest
from openpyxl import Workbook


@pytest.fixture
def sample_excel_bytes():
    """Create a sample Excel file in memory for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TEST_SHEET"
    
    # Header row (row 1)
    ws["A1"] = "Date"
    ws["B1"] = "Value"
    
    # Data rows (rows 2-11, 10 data points)
    for i in range(2, 12):
        ws[f"A{i}"] = datetime(2025, 1, i - 1)
        ws[f"B{i}"] = (i - 1) * 10.5
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.getvalue()


@pytest.fixture
def sample_parser_config():
    """Sample parser configuration."""
    return {
        "parse_config": {
            "series_map": [
                {
                    "internal_series_code": "TEST_SERIES",
                    "unit": "test_unit",
                    "frequency": "D",
                    "sheet": "TEST_SHEET",
                    "header_row": 1,
                    "skip_rows_after_header": 0,
                    "date_col": "A",
                    "value_col": "B",
                    "drop_na": False,
                }
            ]
        }
    }


@pytest.fixture
def sample_normalizer_config():
    """Sample normalizer configuration."""
    return {
        "normalize": {
            "timezone": "UTC",
            "primary_keys": ["obs_time", "internal_series_code"],
        }
    }

