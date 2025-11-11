"""Tests for BCRA Infomondia parser."""

from datetime import datetime

import pytest

from src.infrastructure.plugins.parsers.bcra_infomondia_parser import BcraInfomondiaParser


class TestBcraInfomondiaParser:
    """Tests for BcraInfomondiaParser class."""

    @pytest.fixture
    def parser(self):
        """Create a parser instance."""
        return BcraInfomondiaParser()

    def test_parse_empty_config(self, parser, sample_excel_bytes):
        """Test parsing with empty config."""
        config = {"parse_config": {"series_map": []}}
        result = parser.parse(sample_excel_bytes, config)
        assert result == []

    def test_parse_single_series(self, parser, sample_excel_bytes, sample_parser_config):
        """Test parsing a single series."""
        result = parser.parse(sample_excel_bytes, sample_parser_config)
        
        # Header row is included when skip_rows_after_header is 0
        assert len(result) == 11
        # First item is header, second is first data row
        assert result[1]["internal_series_code"] == "TEST_SERIES"
        assert result[1]["unit"] == "test_unit"
        assert result[1]["frequency"] == "D"
        assert isinstance(result[1]["obs_time"], datetime)
        assert isinstance(result[1]["value"], float)

    def test_parse_with_drop_na(self, parser, sample_excel_bytes):
        """Test parsing with drop_na enabled."""
        config = {
            "parse_config": {
                "series_map": [
                    {
                        "internal_series_code": "TEST_SERIES",
                        "sheet": "TEST_SHEET",
                        "header_row": 1,
                        "skip_rows_after_header": 0,
                        "date_col": "A",
                        "value_col": "B",
                        "drop_na": True,
                    }
                ]
            }
        }
        result = parser.parse(sample_excel_bytes, config)
        # Header row is filtered by drop_na (date is string "Date", not datetime)
        # But drop_na only filters if date OR value is None, header has string "Date" which is not None
        # So header is included, but normalizer will filter it later
        assert len(result) == 11  # Header included (will be filtered by normalizer)

    def test_parse_with_skip_rows(self, parser, sample_excel_bytes):
        """Test parsing with skip_rows_after_header."""
        config = {
            "parse_config": {
                "series_map": [
                    {
                        "internal_series_code": "TEST_SERIES",
                        "sheet": "TEST_SHEET",
                        "header_row": 1,
                        "skip_rows_after_header": 2,
                        "date_col": "A",
                        "value_col": "B",
                        "drop_na": False,
                    }
                ]
            }
        }
        result = parser.parse(sample_excel_bytes, config)
        # Header (1) + skip 2 rows = start at row 4, so 9 rows (4-12, inclusive)
        assert len(result) == 9

    def test_parse_with_series_last_dates_filter(self, parser, sample_excel_bytes, sample_parser_config):
        """Test parsing with series_last_dates filter for incremental updates."""
        series_last_dates = {"TEST_SERIES": datetime(2025, 1, 5)}
        result = parser.parse(sample_excel_bytes, sample_parser_config, series_last_dates)
        
        # Should only include dates after 2025-01-05
        assert len(result) < 10
        for item in result:
            obs_time = item["obs_time"]
            if isinstance(obs_time, datetime):
                assert obs_time > series_last_dates["TEST_SERIES"]

    def test_parse_nonexistent_sheet(self, parser, sample_excel_bytes):
        """Test parsing with non-existent sheet."""
        config = {
            "parse_config": {
                "series_map": [
                    {
                        "internal_series_code": "TEST_SERIES",
                        "sheet": "NONEXISTENT",
                        "header_row": 1,
                        "date_col": "A",
                        "value_col": "B",
                    }
                ]
            }
        }
        result = parser.parse(sample_excel_bytes, config)
        assert result == []

    def test_parse_multiple_series(self, parser, sample_excel_bytes):
        """Test parsing multiple series."""
        config = {
            "parse_config": {
                "series_map": [
                    {
                        "internal_series_code": "SERIES_1",
                        "sheet": "TEST_SHEET",
                        "header_row": 1,
                        "skip_rows_after_header": 0,
                        "date_col": "A",
                        "value_col": "B",
                        "drop_na": False,
                    },
                    {
                        "internal_series_code": "SERIES_2",
                        "sheet": "TEST_SHEET",
                        "header_row": 1,
                        "skip_rows_after_header": 0,
                        "date_col": "A",
                        "value_col": "B",
                        "drop_na": False,
                    },
                ]
            }
        }
        result = parser.parse(sample_excel_bytes, config)
        assert len(result) == 22  # 11 rows (including header) * 2 series

    def test_parse_drop_na_filters_none_values(self, parser):
        """Test that drop_na filters rows with None date or value (line 140)."""
        import io
        from openpyxl import Workbook
        
        # Create Excel with some None values
        wb = Workbook()
        ws = wb.active
        ws.title = "TEST_SHEET"
        ws["A1"] = "Date"
        ws["B1"] = "Value"
        ws["A2"] = datetime(2025, 1, 1)
        ws["B2"] = 100.5
        ws["A3"] = None  # None date
        ws["B3"] = 200.5
        ws["A4"] = datetime(2025, 1, 3)
        ws["B4"] = None  # None value
        ws["A5"] = datetime(2025, 1, 4)
        ws["B5"] = 300.5
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_bytes = excel_file.getvalue()
        
        config = {
            "parse_config": {
                "series_map": [
                    {
                        "internal_series_code": "TEST_SERIES",
                        "sheet": "TEST_SHEET",
                        "header_row": 1,
                        "skip_rows_after_header": 0,
                        "date_col": "A",
                        "value_col": "B",
                        "drop_na": True,  # Should filter None values
                    }
                ]
            }
        }
        
        result = parser.parse(excel_bytes, config)
        # Header row (row 1) is included, then rows 2 and 5 (rows 3 and 4 have None values and are filtered)
        # So we get: header + row 2 + row 5 = 3 items
        assert len(result) == 3
        # First is header (string "Date"), second is row 2, third is row 5
        assert result[1]["value"] == 100.5
        assert result[2]["value"] == 300.5

